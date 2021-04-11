from bs4 import BeautifulSoup
import requests
import time
import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
#from typing import namedtuple


#class Course(NamedTuple):
#    codigoEscola : int
#    nomeEscola : str
#    codigoCurso : str
#    nomeCurso : str 
#    tipoCurso : str

def main():
    excelPath = "excel.xls"
    wb = load_workbook(filename = 'DB.xlsx')
    ws = wb.active
    
    #print("test")
    downloadURL = "https://www.dges.gov.pt/guias/indest.asp"
    data = requests.get(downloadURL)
    #print(data)
    soup = BeautifulSoup(data.text, 'html.parser')
    courses = list(soup.findAll(attrs={'class': 'lin-ce'}))
    #print (len(courses))
    f = open("data.txt", "w")
    sectionHeaders = soup.findAll(attrs={'class': 'box9'})
    for sectHeader in sectionHeaders:
        currentSection = sectHeader.findChildren(attrs={'class': 'lin-area-d2'})
        currentSection2 = sectHeader.findChildren(attrs={'class': 'lin-area-c1'})
        schoolNumber = currentSection2[0].text
        schoolName = currentSection[0].text
        for sibling in sectHeader.findNextSiblings():
            if('class' in sibling.attrs):
                if "box9" in sibling.attrs["class"] or "br" in sibling.name or "nobottomgap" in sibling.attrs["class"]:
                    break
                subDivs = sibling.findChildren()

                courseNumber = subDivs[1].text
                courseURL = "https://www.dges.gov.pt/guias/"+subDivs[2].next["href"]
                courseName = subDivs[3].text
                courseType = subDivs[4].text
                #courseName = sibling.findChildren("a")
                #courseDataURL = courseName[0]["href"]
                #print(courseName[0].getText())
                #print(courseDataURL)

            
                #print("\n"+schoolName)
                #print(courseName)
                #print(courseNumber)
                #print(courseType)
                #print(courseURL)

                courseData = requests.get(courseURL)
                courseSoup = BeautifulSoup(courseData.text, 'html.parser')
                dataCells = list(courseSoup.findAll(attrs={'class': 'tvag'}))

                firstPhaseGPA2019 = 0
                secondPhaseGPA2019 = 0
                secondPhaseGPA2020 = 0
                firstPhaseGPA2020 = 0

                if(len(dataCells) != 0):
                    firstPhaseGPA2020 = dataCells[10].text
                    secondPhaseGPA2020 = dataCells[11].text
                    firstPhaseGPA2019 = dataCells[8].text
                    secondPhaseGPA2019 = dataCells[9].text

                #print(firstPhaseGPA2019)
                #print(secondPhaseGPA2019)
                #print(firstPhaseGPA2020)
                #print(secondPhaseGPA2020)

                infoCursosURL = f"https://infocursos.mec.pt/dges.asp?code={schoolNumber}&codc={courseNumber}"
                txDesemprego = 100
                for row in ws.values:
                    if((row[7] == courseNumber or row[6] == courseName) and (row[1] == schoolNumber or row[3] == schoolNumber)):
                        txDesemprego = row[12]
                print(txDesemprego)

                f.write("\n\n"+schoolName)
                f.write("\n"+schoolNumber)
                f.write("\n"+courseName)
                f.write("\n"+courseNumber)
                f.write("\n"+courseType)
                f.write("\n"+"2019f1: "+str(firstPhaseGPA2019))
                f.write("\n"+"2019f2: "+str(secondPhaseGPA2019))
                f.write("\n"+"2020f1: "+str(firstPhaseGPA2020))
                f.write("\n"+"2020f2: "+str(secondPhaseGPA2020))
                f.write("\n"+str(txDesemprego))
                
                #print (infoCursosURL)

                #for row_num in range(sheet.nrows):
                #    row_value = sheet.row_values(row_num)
                #    print(row_value[0])
                #    print(row_value[6])
                #    if row_value[0] == schoolNumber and row_value[6] == courseNumber:
                #        print (row_value[11])

                #statCourseData = requests.get(infoCursosURL)
                #statCourseSoup = BeautifulSoup(statCourseData.text, 'html.parser')
                #dataCells = list(statCourseSoup.findAll(attrs={'class': 'tvag'}))
    f.close()
            


    
if __name__ == '__main__':
    main()